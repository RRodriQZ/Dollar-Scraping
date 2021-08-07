from openpyxl import load_workbook, Workbook
from configparser import ConfigParser
from log.logger import Log
import os


# GLOBAL VALUES #
config = ConfigParser()
config.read("sheets/resources/config.ini")

REPORT_FILE = config["File"]["excel_file"]
HEAD_COLUMN = config["File"]["head"]
logger = Log().get_logger(__name__)


def check_exists_excel_file() -> None:
    try:
        print(f"========================= STARTING THE SCRIPT =========================\n")
        if os.path.exists(REPORT_FILE):
            print(f'=========== Updating ... the file "{REPORT_FILE}" ===========\n')
            logger.info(f"========================= STARTING SCRAPING =========================")
            logger.info(f'=========== Updating ... the file "{REPORT_FILE}" ===========')

        else:
            print(f'=========== File does not exist: "{REPORT_FILE}" ===========')

            head_column = HEAD_COLUMN.split(",")
            titles = tuple(head_column)

            wb = Workbook()
            wb.save(REPORT_FILE)

            sheet = wb.active
            sheet.append(titles)

            wb.save(REPORT_FILE)

            print(f'Generated a new file: "{REPORT_FILE}"')
            logger.info(f'=========== File does not exist: "{REPORT_FILE}" ===========')
            logger.info(f'Generated a new file: "{REPORT_FILE}"')

    except Exception as e:
        print(f'ERROR in the Excel Generation "{REPORT_FILE}", error: "{e}"')
        logger.error(f'ERROR in the Excel Generation "{REPORT_FILE}", error: "{e}"')


def update_sheet_row(data: tuple) -> None:
    try:
        wb = load_workbook(REPORT_FILE)

        sheet = wb.active

        sheet.append(data)

        wb.save(REPORT_FILE)

        print(f'=========== "{REPORT_FILE}" was correctly updated with the new information ===========')
        logger.info(f'=========== "{REPORT_FILE}" was correctly updated with the new information ===========')

    except Exception as e:
        print(f'ERROR in the excel update, error: "{e}"')
        logger.error(f'ERROR in the excel update, error: "{e}"')
